"""
=============================================================================
TỐI ƯU HÓA ĐÈN GIAO THÔNG – TRỤC ĐIỆN BIÊN PHỦ / VÕ THỊ SÁU / LÝ CHÍNH THẮNG
Thuật toán: NSGA-II + Local Search (Tích hợp AHP)
Cải tiến: Sử dụng CHU KỲ CHUNG (C) cho toàn bộ mạng lưới.
Bảo toàn 100% công thức toán học (Tích hợp chống tràn bến L_max).
=============================================================================
"""
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import warnings
warnings.filterwarnings('ignore')
import matplotlib.patches as mpatches
import time
import random
import os

os.makedirs('outputs', exist_ok=True)

plt.rcParams['font.family'] = 'DejaVu Sans'
plt.rcParams['axes.unicode_minus'] = False
plt.style.use('seaborn-v0_8-whitegrid')
COLORS = ['#2196F3', '#FF5722', '#4CAF50', '#FF9800', '#9C27B0','#00BCD4', '#F44336', '#8BC34A', '#3F51B5', '#FFC107']

# ===================== BIẾN TOÀN CỤC CHỨA DỮ LIỆU =====================
df_nodes = pd.DataFrame()
DISTANCES = []
L_MAX_LIST = [] 
BASELINE_C = []
BASELINE_G1 = []
BASELINE_G2 = []
BASELINE_OFF = []
BASELINE_F1 = []
BASELINE_F2 = []
BASELINE_F3 = []

C_MIN, C_MAX = 85, 95 
G_MIN, G_MAX = 15, 80
OFF_MIN, OFF_MAX = 0, 149

# ===================== HÀM ĐỌC DỮ LIỆU TỪ EXCEL =====================
def load_data_from_excel():
    global df_nodes, DISTANCES, BASELINE_C, BASELINE_G1, BASELINE_G2, BASELINE_OFF
    global BASELINE_F1, BASELINE_F2, BASELINE_F3, L_MAX_LIST
    global C_MIN, C_MAX
    
    print("[+] Đang nạp dữ liệu từ file Excel...")
    try:
        try:
            df_nut = pd.read_excel('du_lieu_nut_mang_luoi.xlsx', header=3)
            if not any('q_total' in str(c) for c in df_nut.columns):
                df_nut = pd.read_excel('du_lieu_nut_mang_luoi.xlsx')
        except:
            df_nut = pd.read_excel('du_lieu_nut_mang_luoi.xlsx')
            
        col_q = next((c for c in df_nut.columns if 'q_total' in str(c) or c == 'q'), df_nut.columns[2])
        col_s = next((c for c in df_nut.columns if 'S' in str(c)), df_nut.columns[3])
        col_name = next((c for c in df_nut.columns if 'Tên' in str(c) or 'name' in str(c)), df_nut.columns[1])
        col_L = next((c for c in df_nut.columns if 'L(' in str(c) or c == 'L'), df_nut.columns[4])
        col_v = next((c for c in df_nut.columns if 'v' in str(c) and 'km' in str(c).lower()), df_nut.columns[5])
        col_qb = next((c for c in df_nut.columns if 'qb' in str(c)), df_nut.columns[6])
        
        col_g1 = next(c for c in df_nut.columns if 'Pha1' in str(c) or 'g1' in str(c))
        col_g2 = next(c for c in df_nut.columns if 'Pha2' in str(c) or 'g2' in str(c))
        col_off = next(c for c in df_nut.columns if 'Offset' in str(c) or 'o_bl' in str(c))
        n_nodes = len(df_nut)
        
        nodes_dict = {
            'k': list(range(1, n_nodes + 1)),
            'name': df_nut[col_name].astype(str).tolist(),
            'q': df_nut[col_q].fillna(1500).astype(float).tolist(),
            'S': df_nut[col_s].fillna(3600).astype(float).tolist(),
            'L': df_nut[col_L].fillna(3).astype(float).tolist(),
            'v': df_nut[col_v].fillna(15.0).astype(float).tolist(),
            'qb': df_nut[col_qb].fillna(15).astype(float).tolist()
        }
        df_nodes = pd.DataFrame(nodes_dict)
        
        df_lk = pd.read_excel('Khoang_Cach_mang_luoi.xlsx')
        col_tu_nut = next((c for c in df_lk.columns if 'Tu nut' in str(c) or 'Tu_Nut' in str(c)), df_lk.columns[0])
        col_kc = next((c for c in df_lk.columns if 'Khoang_cach' in str(c) or 'Khoang_Cach' in str(c)), df_lk.columns[2])
        
        DISTANCES = [0.0] * n_nodes
        for idx, row in df_lk.iterrows():
            tu = int(row.get(col_tu_nut, idx + 1))
            if tu - 1 < n_nodes:
                DISTANCES[tu - 1] = float(row.get(col_kc, 200.0))
                
        # ================= BỔ SUNG LOGIC TÍNH L_MAX =================
        L_XE = 6.0     
        N_LAN = 3      
        ALPHA = 0.85   
        L_MAX_LIST = []
        for d in DISTANCES:
            if d > 0:
                # Công thức: L_max = (Khoảng cách / Chiều dài xe) * Số làn * Hệ số an toàn
                max_veh = round((d / L_XE) * N_LAN * ALPHA)
                L_MAX_LIST.append(max_veh)
            else:
                L_MAX_LIST.append(9999) 
        BASELINE_C = [90] * n_nodes 
        BASELINE_G1 = df_nut[col_g1].fillna(50).astype(int).tolist()
        BASELINE_G2 = df_nut[col_g2].fillna(34).astype(int).tolist()
        BASELINE_OFF = df_nut[col_off].fillna(0).astype(int).tolist()

        max_L = max([float(l) for l in df_nodes['L']])
        absolute_c_min = max(60, int(2 * G_MIN + 2 * max_L))
        C_MIN = max(absolute_c_min, 85)
        C_MAX = max(C_MIN + 5, 95)

        BASELINE_F1, BASELINE_F2, BASELINE_F3 = [], [], []
        for i in range(n_nodes):
            q, S, L, v, qb = df_nodes['q'][i], df_nodes['S'][i], df_nodes['L'][i], df_nodes['v'][i], df_nodes['qb'][i]
            c, g1, g2, off = BASELINE_C[i], BASELINE_G1[i], BASELINE_G2[i], BASELINE_OFF[i]
            q1, q2 = q * 0.6, q * 0.4
            
            d1 = calc_uniform_delay(q1, S*0.6, g1, c) + calc_incremental_delay(q1, S*0.6, g1, c) + calc_residual_delay(q1, qb*0.6)
            d2 = calc_uniform_delay(q2, S*0.4, g2, c) + calc_incremental_delay(q2, S*0.4, g2, c) + calc_residual_delay(q2, qb*0.4)
            BASELINE_F1.append((d1*q1 + d2*q2)/3600)
            
            lq1 = calc_lq1_uniform(q1, S*0.6, g1, c) + calc_lq2_random(q1, S*0.6, g1, c)
            lq2 = calc_lq1_uniform(q2, S*0.4, g2, c) + calc_lq2_random(q2, S*0.4, g2, c)
            BASELINE_F2.append(lq1 + lq2) 
            
            t_travel = DISTANCES[i] / (v * 1000/3600 + 1e-6)
            gamma = calc_gamma_wave(off, t_travel, c)
            ns1 = calc_node_stops(q1, S*0.6, g1, c, gamma)
            ns2 = calc_node_stops(q2, S*0.4, g2, c, gamma)
            BASELINE_F3.append(ns1 + ns2)
            
        print(f"[+] Load thành công {n_nodes} nút từ Excel. Giới hạn chu kỳ: [{C_MIN}s - {C_MAX}s].")
        
    except Exception as e:
        print(f"[!] Lỗi đọc Excel: {e}. Vui lòng kiểm tra lại cấu trúc file.")
        exit()

# ===================== HỆ THỐNG CÔNG THỨC TOÁN HỌC CHUẨN =====================
def calc_uniform_delay(q, S, g, C):
    if g <= 0 or S <= 0 or C <= 0: return 9999.0
    lam = g / C
    x = (q * C) / (S * g)
    num = C * ((1 - lam) ** 2)
    den = 2 * (1 - min(1.0, x) * lam)
    return max(0.0, num / den)

def calc_incremental_delay(q, S, g, C, T_period=1):
    if g <= 0 or S <= 0: return 0.0
    lam = g / C
    x = (q * C) / (S * g)
    c_cap = S * lam
    term = (8 * 0.5 * 1.0 * x) / (c_cap * T_period + 1e-6)
    d2 = 900 * T_period * ((x - 1) + np.sqrt(max(0.0, (x - 1) ** 2 + term)))
    return max(0.0, d2)

def calc_residual_delay(q, qb, t_accum=30, T_period=1.0):
    if q <= 0: return 0.0
    return (3600.0 * qb * t_accum) / (q * T_period + 1e-6)

def calc_lq1_uniform(q, S, g, C):
    if C <= 0 or g <= 0: return 0.0
    lam = g / C
    x = (q * C) / (S * g)
    x = min(0.95, x)
    r = max(0.0, C - g)
    num = (q / 3600.0) * (r ** 2)
    den = 2 * C * (1 - x * lam + 1e-6)
    return max(0.0, num / den)
def calc_lq2_random(q, S, g, C, T_period=1):
    if g <= 0 or S <= 0: return 0.0
    lam = g / C
    x = (q * C) / (S * g)
    if x <= 0.5: return 0.0
    c_cap = S * lam
    term = ((x - 1) ** 2) + (8 * 0.5 * 1.0 * x) / (c_cap * T_period + 1e-6)
    lq2 = 0.25 * c_cap * T_period * ((x - 1) + np.sqrt(max(0.0, term)))
    
    return max(0.0, lq2)

def calc_gamma_wave(offset_k, t_travel, C, beta=0.3):
    diff = abs(offset_k - t_travel) % C
    if diff > C / 2: diff = C - diff
    return max(0.05, 1.0 - beta * np.exp(-diff / (C / 4 + 1e-6)))

def calc_node_stops(q, S, g, C, gamma):
    if C <= 0 or q <= 0: return 0.0
    x = (q * C) / (S * g + 1e-6)
    term_red = (C - g) / C
    term_sat = 1.0 / (1.0 - min(0.95, x) + 1e-6)
    return q * term_red * term_sat * gamma

def evaluate_individual(individual):
    n = len(df_nodes)
    G1 = individual[:n]
    OFF = individual[n:2*n]
    c = individual[-1]

    f1_total, f2_total, f3_total = 0.0, 0.0, 0.0

    # Ngưỡng bão hòa tối đa cho phép (Ràng buộc 4)
    X_MAX = 0.95 

    for i in range(n):
        q, S, L, v, qb = df_nodes['q'][i], df_nodes['S'][i], df_nodes['L'][i], df_nodes['v'][i], df_nodes['qb'][i]
        g1, off_k = G1[i], OFF[i]
        
        g2 = max(G_MIN, c - g1 - 2*int(L))
        q1, q2 = q * 0.6, q * 0.4
        S1, S2 = S * 0.6, S * 0.4

        # F1: TỔNG ĐỘ TRỄ
        d1 = calc_uniform_delay(q1, S1, g1, c) + calc_incremental_delay(q1, S1, g1, c) + calc_residual_delay(q1, qb*0.6)
        d2_delay = calc_uniform_delay(q2, S2, g2, c) + calc_incremental_delay(q2, S2, g2, c) + calc_residual_delay(q2, qb*0.4)
        f1_total += (d1 * q1 + d2_delay * q2) / 3600.0

        # ================= SỬA LỖI 4: RÀNG BUỘC HỆ SỐ BÃO HÒA x_max =================
        x1 = (q1 * c) / (S1 * g1 + 1e-6)
        x2 = (q2 * c) / (S2 * g2 + 1e-6)
        
        penalty_x = 0
        if x1 > X_MAX:
            penalty_x += (x1 - X_MAX) * 10000 # Phạt rất nặng nếu x vượt 0.95
        if x2 > X_MAX:
            penalty_x += (x2 - X_MAX) * 10000
            
        # Áp dụng phạt bão hòa vào các hàm mục tiêu
        f1_total += penalty_x
        f2_total += penalty_x
        f3_total += penalty_x
        # =============================================================================
        lq1 = calc_lq1_uniform(q1, S1, g1, c) + calc_lq2_random(q1, S1, g1, c)
        lq2_queue = calc_lq1_uniform(q2, S2, g2, c) + calc_lq2_random(q2, S2, g2, c)
        
        penalty_lq = 0
        
        if lq1 > L_MAX_LIST[i]:
            penalty_lq += (lq1 - L_MAX_LIST[i]) * 1000
        
        if lq2_queue > L_MAX_LIST[i]:
            penalty_lq += (lq2_queue - L_MAX_LIST[i]) * 1000
        
        f1_total += penalty_lq
        f2_total += (lq1 + lq2_queue) + penalty_lq
        f3_total += penalty_lq
        # =============================================================================

        # F3: SỐ LƯỢT DỪNG
        t_travel = DISTANCES[i] / (v * 1000/3600 + 1e-6)
        gamma = calc_gamma_wave(off_k, t_travel, c)
        ns1 = calc_node_stops(q1, S1, g1, c, gamma)
        ns2 = calc_node_stops(q2, S2, g2, c, gamma)
        f3_total += (ns1 + ns2)

    return f1_total, f2_total, f3_total
# ===================== CẤU TRÚC INDIVIDUAL & NSGA-II =====================
def repair_individual(ind):
    n = len(df_nodes)
    ind = list(ind)
    c = int(np.clip(ind[-1], C_MIN, C_MAX))
    ind[-1] = c
    
    for i in range(n):
        L = int(df_nodes['L'][i])
        g1 = int(np.clip(ind[i], G_MIN, G_MAX))
        if c - g1 - 2*L < G_MIN:
            g1 = int(np.clip(c - G_MIN - 2*L, G_MIN, G_MAX))
        ind[i] = g1
        ind[n+i] = int(ind[n+i] % c) if c > 0 else 0
    return ind

def create_individual(force_c=None):
    n = len(df_nodes)
    c = force_c if force_c else random.randint(C_MIN, C_MAX)
    G1, OFF = [], []
    for i in range(n):
        L = int(df_nodes['L'][i])
        max_g = max(G_MIN, c - G_MIN - 2*L)
        G1.append(random.randint(G_MIN, max(G_MIN, max_g)))
        OFF.append(random.randint(OFF_MIN, c-1))
    return repair_individual(G1 + OFF + [c])

def crossover(p1, p2):
    n = len(p1)
    c1, c2 = list(p1), list(p2)
    for i in range(n):
        if random.random() < 0.5:
            c1[i], c2[i] = c2[i], c1[i]
    return repair_individual(c1), repair_individual(c2)

def mutate(ind, pm=0.2):
    n = len(df_nodes)
    ind = list(ind)
    if random.random() < 0.5: 
        for _ in range(random.randint(2, 6)):
            k = random.randint(0, 2*n) 
            if k < n: 
                ind[k] += random.choice([-5, -3, 3, 5])
            elif k < 2*n: 
                ind[k] += random.choice([-10, 10])
            else: 
                ind[k] += random.choice([-8, 0, 8])
    return repair_individual(ind)

def dominates(a, b):
    return all(ai <= bi for ai, bi in zip(a, b)) and any(ai < bi for ai, bi in zip(a, b))

def fast_non_dominated_sort(pop_fit):
    n = len(pop_fit)
    S_dom, n_dom, rank = [[] for _ in range(n)], [0]*n, [0]*n
    fronts = [[]]
    for p in range(n):
        for q in range(n):
            if p == q: continue
            if dominates(pop_fit[p], pop_fit[q]): S_dom[p].append(q)
            elif dominates(pop_fit[q], pop_fit[p]): n_dom[p] += 1
        if n_dom[p] == 0:
            rank[p] = 1
            fronts[0].append(p)
    i = 0
    while fronts[i]:
        next_front = []
        for p in fronts[i]:
            for q in S_dom[p]:
                n_dom[q] -= 1
                if n_dom[q] == 0:
                    rank[q] = i + 2
                    next_front.append(q)
        i += 1
        fronts.append(next_front)
    return fronts[:-1], rank

def crowding_distance(pop_fit, front):
    n = len(front)
    if n == 0: return []
    dist = [0.0]*n
    for obj in range(len(pop_fit[0])):
        sorted_idx = sorted(range(n), key=lambda i: pop_fit[front[i]][obj])
        dist[sorted_idx[0]] = dist[sorted_idx[-1]] = float('inf')
        f_min, f_max = pop_fit[front[sorted_idx[0]]][obj], pop_fit[front[sorted_idx[-1]]][obj]
        if f_max == f_min: continue
        for k in range(1, n-1):
            dist[sorted_idx[k]] += (pop_fit[front[sorted_idx[k+1]]][obj] - pop_fit[front[sorted_idx[k-1]]][obj]) / (f_max - f_min)
    return dist

def select_tournament(pop, pop_fit, ranks, crowding, k=3):
    candidates = random.sample(range(len(pop)), k)
    best = candidates[0]
    for c in candidates[1:]:
        if ranks[c] < ranks[best] or (ranks[c] == ranks[best] and crowding[c] > crowding[best]):
            best = c
    return pop[best]

def nsga2(pop_size=120, n_gen=80, seed=42):
    random.seed(seed)
    np.random.seed(seed)
    print(f"\n[NSGA-II] Khởi tạo quần thể {pop_size} cá thể (Chu kỳ chung), {n_gen} thế hệ...")

    pop = [repair_individual(BASELINE_G1 + BASELINE_OFF + [90])]
    pop += [create_individual(force_c=90) for _ in range(pop_size // 3)]
    pop += [create_individual() for _ in range(pop_size - 1 - pop_size // 3)]

    pop_fit = [evaluate_individual(ind) for ind in pop]
    fronts, ranks = fast_non_dominated_sort(pop_fit)
    crowding = [0.0]*pop_size
    for front in fronts:
        cd = crowding_distance(pop_fit, front)
        for j, idx in enumerate(front): crowding[idx] = cd[j]

    history = {'gen': [], 'f1': [], 'f2': [], 'f3': []}

    for gen in range(n_gen):
        offspring = []
        while len(offspring) < pop_size:
            p1 = select_tournament(pop, pop_fit, ranks, crowding)
            p2 = select_tournament(pop, pop_fit, ranks, crowding)
            c1, c2 = crossover(p1, p2) 
            offspring.extend([mutate(c1), mutate(c2)])
        offspring = offspring[:pop_size]

        combined = pop + offspring
        combined_fit = pop_fit + [evaluate_individual(ind) for ind in offspring]
        fronts_c, ranks_c = fast_non_dominated_sort(combined_fit)
        
        crowding_c = [0.0]*len(combined)
        for front in fronts_c:
            cd = crowding_distance(combined_fit, front)
            for j, idx in enumerate(front): crowding_c[idx] = cd[j]

        pop, pop_fit, ranks, crowding = [], [], [], []
        for front in fronts_c:
            if len(pop) + len(front) <= pop_size:
                for idx in front:
                    pop.append(combined[idx]); pop_fit.append(combined_fit[idx])
                    ranks.append(ranks_c[idx]); crowding.append(crowding_c[idx])
            else:
                rem = pop_size - len(pop)
                sorted_front = sorted(front, key=lambda i: -crowding_c[i])
                for idx in sorted_front[:rem]:
                    pop.append(combined[idx]); pop_fit.append(combined_fit[idx])
                    ranks.append(ranks_c[idx]); crowding.append(crowding_c[idx])
                break

        pareto_fits = [pop_fit[i] for i in range(len(pop)) if ranks[i] == 1]
        if pareto_fits:
            history['gen'].append(gen+1)
            history['f1'].append(np.mean([f[0] for f in pareto_fits]))
            history['f2'].append(np.mean([f[1] for f in pareto_fits]))
            history['f3'].append(np.mean([f[2] for f in pareto_fits]))

        if (gen+1) % 10 == 0:
            print(f"    Gen {gen+1}/{n_gen} | Pareto: {len(pareto_fits)} | f1_mean={history['f1'][-1]:.1f} | f2_mean={history['f2'][-1]:.1f} | f3_mean={history['f3'][-1]:.0f}")

    pareto_idx = [i for i in range(len(pop)) if ranks[i] == 1]
    return [pop[i] for i in pareto_idx], [pop_fit[i] for i in pareto_idx], history

# ===================== AHP & LOCAL SEARCH =====================
def ahp_pairwise_matrix():
    return np.array([[1.0, 3.0, 4.0], [1/3, 1.0,2.0], [1/4, 1/2, 1.0]])

def compute_ahp_weights(A):
    n = A.shape[0]
    geo_means = np.array([np.prod(A[i, :])**(1/n) for i in range(n)])
    weights = geo_means / geo_means.sum()
    lambda_max = np.mean((A @ weights) / weights)
    CI = (lambda_max - n) / (n - 1)
    CR = CI / 0.58
    return weights, lambda_max, CI, CR, A

def weighted_sum_score(fit, weights, f_min, f_max):
    score = 0
    for j in range(3):
        norm = (fit[j] - f_min[j]) / (f_max[j] - f_min[j]) if f_max[j] > f_min[j] else 0
        score += weights[j] * norm
    return score

def local_search(pareto_pop, pareto_fits, weights, n_iter=50):
    print(f"\n[Local Search] Cường độ cao cho {len(pareto_pop)} cá thể Pareto...")
    all_fits = np.array(pareto_fits)
    f_min, f_max = all_fits.min(axis=0), all_fits.max(axis=0)

    scores = [weighted_sum_score(f, weights, f_min, f_max) for f in pareto_fits]
    top_k = max(1, int(len(pareto_pop)*0.3)) 
    top_idx = sorted(range(len(scores)), key=lambda i: scores[i])[:top_k]

    best_ind, best_fit, best_score = pareto_pop[top_idx[0]], pareto_fits[top_idx[0]], scores[top_idx[0]]
    improved_count = 0
    ls_history = [best_score]

    for idx in top_idx:
        ind, fit, score = list(pareto_pop[idx]), list(pareto_fits[idx]), scores[idx]
        for _ in range(n_iter):
            neighbor = list(ind)
            n_var = len(df_nodes)
            for k in random.sample(range(2*n_var + 1), random.randint(2, 5)):
                if k < n_var: 
                    neighbor[k] += random.choice([-2, 0, 2])
                elif k < 2*n_var: 
                    neighbor[k] += random.choice([-8, 8])
                else: 
                    neighbor[k] += random.choice([-5, 0, 5])
            
            neighbor = repair_individual(neighbor)
            new_fit = evaluate_individual(neighbor)
            new_score = weighted_sum_score(new_fit, weights, f_min, f_max)

            if new_score < score:
                ind, fit, score = neighbor, list(new_fit), new_score
                improved_count += 1
                ls_history.append(score)

        if score < best_score:
            best_ind, best_fit, best_score = ind, fit, score

    print(f"  > Tìm được {improved_count} cải thiện cục bộ.")
    return best_ind, best_fit, ls_history

# ===================== TRỰC QUAN HÓA =====================
def plot_input_data():
    fig, axes = plt.subplots(2, 2, figsize=(16, 10))
    fig.suptitle('DỮ LIỆU ĐẦU VÀO MẠNG LƯỚI GIAO THÔNG', fontsize=14, fontweight='bold', y=1.01)
    nodes_x = list(range(1, len(df_nodes)+1))
    labels_short = [f'N{k}' for k in range(1, len(df_nodes)+1)]

    ax = axes[0,0]
    ax.bar(nodes_x, df_nodes['q'], color=COLORS[0], alpha=0.8, edgecolor='white', label='Lưu lượng xe (q)')
    ax.set_title('Lưu lượng xe (q) – xe/giờ', fontweight='bold')
    ax.set_xticks(nodes_x); ax.set_xticklabels(labels_short, rotation=45, fontsize=7)
    ax.axhline(df_nodes['q'].mean(), color='red', linestyle='--', alpha=0.7, label='Lưu lượng trung bình')
    ax.legend(loc='upper right')

    ax = axes[0,1]
    ax.bar(nodes_x, df_nodes['S'], color=COLORS[2], alpha=0.8, edgecolor='white', label='Lưu lượng bão hòa (S)')
    ax.set_title('Lưu lượng bão hòa (S) – xe/giờ xanh', fontweight='bold')
    ax.set_xticks(nodes_x); ax.set_xticklabels(labels_short, rotation=45, fontsize=7)
    ax.legend(loc='upper right')

    ax = axes[1,0]
    x_ratio = [df_nodes['q'][i]*100/(df_nodes['S'][i]*BASELINE_G1[i]) for i in range(len(df_nodes))]
    colors_x = [COLORS[1] if x > 0.8 else COLORS[0] for x in x_ratio]
    ax.bar(nodes_x, x_ratio, color=colors_x, alpha=0.8, edgecolor='white')
    ax.axhline(0.9, color='red', linestyle='--', label='Ngưỡng bão hòa (0.9)')
    ax.axhline(0.75, color='orange', linestyle='--', label='Ngưỡng cảnh báo (0.75)')
    ax.set_title('Hệ số bão hòa x (Baseline)', fontweight='bold')
    ax.set_xticks(nodes_x); ax.set_xticklabels(labels_short, rotation=45, fontsize=7)
    
    patch_high = mpatches.Patch(color=COLORS[1], label='x > 0.8 (Cảnh báo cao)')
    patch_normal = mpatches.Patch(color=COLORS[0], label='x ≤ 0.8 (Bình thường)')
    ax.legend(handles=[patch_high, patch_normal, ax.lines[0], ax.lines[1]], loc='upper left', fontsize=9)

    ax = axes[1,1]
    ax2 = ax.twinx()
    ax.plot(nodes_x, df_nodes['v'], 'bo-', label='Tốc độ (km/h)')
    ax2.plot(nodes_x[:-1], DISTANCES[:-1], 'r^--', label='Khoảng cách (m)')
    ax.set_title('Tốc độ & Khoảng cách', fontweight='bold')
    ax.set_xticks(nodes_x); ax.set_xticklabels(labels_short, rotation=45, fontsize=7)
    
    lines_1, labels_1 = ax.get_legend_handles_labels()
    lines_2, labels_2 = ax2.get_legend_handles_labels()
    ax2.legend(lines_1 + lines_2, labels_1 + labels_2, loc='upper left')

    plt.tight_layout()
    plt.savefig('outputs/01_input_data.png', dpi=150)
    plt.close()

def run_ahp_analysis():
    print("\n" + "="*60 + "\nPHÂN TÍCH AHP – XÁC ĐỊNH TRỌNG SỐ MỤC TIÊU\n" + "="*60)
    A = ahp_pairwise_matrix()
    weights, lambda_max, CI, CR, _ = compute_ahp_weights(A)
    print(pd.DataFrame(A, index=['f1', 'f2', 'f3'], columns=['f1', 'f2', 'f3']).round(3).to_string())
    print(f"\nCR = {CR:.4f} {'✓ Nhất quán' if CR < 0.1 else '✗ Cần xem lại'}")
    return weights, A, lambda_max, CI, CR

def plot_ahp(A, weights, lambda_max, CI, CR):
    fig, axes = plt.subplots(1, 3, figsize=(16, 5))
    fig.suptitle('PHÂN TÍCH AHP – XÁC ĐỊNH TRỌNG SỐ MỤC TIÊU', fontsize=14, fontweight='bold')
    criteria = ['f1 (Delay)', 'f2 (Queue)', 'f3 (Stops)']
    
    im = axes[0].imshow(A, cmap='Blues', aspect='auto')
    axes[0].set_xticks(range(3)); axes[0].set_yticks(range(3))
    axes[0].set_xticklabels(criteria); axes[0].set_yticklabels(criteria)
    axes[0].set_title('Ma trận so sánh cặp', fontweight='bold')
    for i in range(3):
        for j in range(3): 
            axes[0].text(j, i, f'{A[i,j]:.2f}', ha='center', va='center', color='white' if A[i,j]>2 else 'black')
    
    axes[1].bar(criteria, weights*100, color=COLORS[:3], alpha=0.85)
    axes[1].set_title('Trọng số AHP (%)', fontweight='bold')
    
    axes[2].axis('off')
    info = [f"λ_max = {lambda_max:.4f}", f"CI = {CI:.4f}", f"CR = {CR:.4f}"]
    for idx, line in enumerate(info): 
        axes[2].text(0.1, 0.8 - idx*0.1, line, fontsize=12)
    plt.tight_layout()
    plt.savefig('outputs/02_ahp_analysis.png')
    plt.close()

def plot_algorithm_efficiency(history, ls_history):
    fig, axes = plt.subplots(1, 2, figsize=(16, 5))
    fig.suptitle('HIỆU QUẢ CẢI THIỆN CỦA MÔ HÌNH TOÁN TỐI ƯU', fontsize=14, fontweight='bold')

    ax1 = axes[0]
    ax1.plot(history['gen'], history['f1'], color=COLORS[0], lw=2.5, marker='o', ms=3, label='Độ trễ TB (f1)')
    ax1.set_title('Sự hội tụ của NSGA-II theo Thế hệ', fontweight='bold')
    ax1.set_xlabel('Thế hệ (Generation)')
    ax1.set_ylabel('Giá trị hàm mục tiêu')
    ax1.legend()

    ax2 = axes[1]
    ax2.plot(range(len(ls_history)), ls_history, color=COLORS[1], lw=2, marker='^', ms=5)
    ax2.set_title('Quá trình tinh chỉnh cục bộ (Local Search)', fontweight='bold')
    ax2.set_xlabel('Số bước cải thiện tìm thấy')
    ax2.set_ylabel('Điểm AHP (Càng thấp càng tốt)')

    plt.tight_layout()
    plt.savefig('outputs/03_algorithm_efficiency.png')
    plt.close()

def plot_pareto_front(pareto_fits, best_fit):
    fig = plt.figure(figsize=(16, 6))
    pf = np.array(pareto_fits)
    f1, f2, f3 = pf[:, 0], pf[:, 1], pf[:, 2]

    ax1 = fig.add_subplot(131, projection='3d')
    sc = ax1.scatter(f1, f2, f3, c=f1, cmap='viridis', s=40, alpha=0.7)
    ax1.scatter([best_fit[0]], [best_fit[1]], [best_fit[2]], c='red', s=200, marker='*', label='Best')
    ax1.set_title('3D Pareto Front')

    ax2 = fig.add_subplot(132)
    ax2.scatter(f1, f2, c=f3, cmap='plasma', s=50, alpha=0.7)
    ax2.scatter(best_fit[0], best_fit[1], c='red', s=200, marker='*')
    ax2.set_title('f1 vs f2')

    ax3 = fig.add_subplot(133)
    ax3.scatter(f1, f3, c=f2, cmap='cool', s=50, alpha=0.7)
    ax3.scatter(best_fit[0], best_fit[2], c='red', s=200, marker='*')
    ax3.set_title('f1 vs f3')

    plt.tight_layout()
    plt.savefig('outputs/04_pareto_front.png')
    plt.close()

def plot_signal_plan(best_ind):
    n = len(df_nodes)
    G1_opt = best_ind[:n]
    OFF_opt = best_ind[n:2*n]
    c_opt = best_ind[-1]
    G2_opt = [max(G_MIN, c_opt - G1_opt[i] - 2*int(df_nodes['L'][i])) for i in range(n)]

    fig, ax1 = plt.subplots(figsize=(16, 8))
    nodes_x = np.arange(1, n+1)
    labels_short = [f'nút {k}' for k in range(1, n+1)]
    width = 0.55

    ax1.bar(nodes_x, G1_opt, width, color='#90CAF9', label='Pha 1', edgecolor='white')
    ax1.bar(nodes_x, G2_opt, width, bottom=G1_opt, color='#80CBC4', label='Pha 2', edgecolor='white')
    
    ax1.set_xlabel('Các nút giao thông', fontsize=11, fontweight='bold')
    ax1.set_ylabel('Thời gian xanh hiệu dụng (g) (giây)', fontsize=11, fontweight='bold')
    ax1.set_title(f'PHÂN BỔ THỜI GIAN XANH VÀ OFFSET TỐI ƯU (CHU KỲ CHUNG C = {c_opt}s)', fontweight='bold', fontsize=14)
    ax1.set_xticks(nodes_x)
    ax1.set_xticklabels(labels_short, rotation=45, fontsize=10)
    ax1.grid(axis='y', linestyle='--', alpha=0.6)

    ax2 = ax1.twinx()
    ax2.plot(nodes_x, OFF_opt, color='#FF5722', marker='D', linewidth=2.5, markersize=7, label='Độ lệch pha - Offset')
    ax2.set_ylabel('Offset (giây)', fontweight='bold', color='#FF5722', fontsize=11)
    ax2.tick_params(axis='y', labelcolor='#FF5722')

    lines_1, labels_1 = ax1.get_legend_handles_labels()
    lines_2, labels_2 = ax2.get_legend_handles_labels()
    ax1.legend(lines_1 + lines_2, labels_1 + labels_2, loc='upper center', bbox_to_anchor=(0.5, -0.1), ncol=3, fontsize=11)

    plt.tight_layout()
    plt.savefig('outputs/05_signal_plan.png')
    plt.close()

def plot_comparison(best_ind, best_fit):
    n = len(df_nodes)
    G1_opt, OFF_opt, c = best_ind[:n], best_ind[n:2*n], best_ind[-1]
    f1_opt, f2_opt, f3_opt = [], [], []
    for i in range(n):
        q, S, L, v, qb = df_nodes['q'][i], df_nodes['S'][i], df_nodes['L'][i], df_nodes['v'][i], df_nodes['qb'][i]
        g1, off = G1_opt[i], OFF_opt[i]
        g2 = max(G_MIN, c - g1 - 2*L)
        q1, q2 = q*0.6, q*0.4
        
        d1 = calc_uniform_delay(q1, S*0.6, g1, c) + calc_incremental_delay(q1, S*0.6, g1, c) + calc_residual_delay(q1, qb*0.6)
        d2 = calc_uniform_delay(q2, S*0.4, g2, c) + calc_incremental_delay(q2, S*0.4, g2, c) + calc_residual_delay(q2, qb*0.4)
        f1_opt.append((d1*q1 + d2*q2)/3600)
        
        # BỔ SUNG: Phần hiển thị kết quả vẫn giữ nguyên để báo cáo được giá trị thật (chưa cộng hình phạt)
        lq1 = calc_lq1_uniform(q1, S*0.6, g1, c) + calc_lq2_random(q1, S*0.6, g1, c)
        lq2 = calc_lq1_uniform(q2, S*0.4, g2, c) + calc_lq2_random(q2, S*0.4, g2, c)
        f2_opt.append(lq1 + lq2)
        
        t_travel = DISTANCES[i] / (v * 1000/3600 + 1e-6)
        gamma = calc_gamma_wave(off, t_travel, c)
        f3_opt.append(calc_node_stops(q1, S*0.6, g1, c, gamma) + calc_node_stops(q2, S*0.4, g2, c, gamma))
        
    pct_f1 = [(bl-opt)/bl*100 if bl>0 else 0 for bl, opt in zip(BASELINE_F1, f1_opt)]
    pct_f2 = [(bl-opt)/bl*100 if bl>0 else 0 for bl, opt in zip(BASELINE_F2, f2_opt)]
    pct_f3 = [(bl-opt)/bl*100 if bl>0 else 0 for bl, opt in zip(BASELINE_F3, f3_opt)]
    
    fig, axes = plt.subplots(3, 1, figsize=(18, 15))
    nodes_x = np.arange(1, n+1)
    labels_short = [f'N{k}' for k in range(1, n+1)]
    
    for i, (ax, bl, opt, pct, title) in enumerate(zip(axes, 
                                                      [BASELINE_F1, BASELINE_F2, BASELINE_F3], 
                                                      [f1_opt, f2_opt, f3_opt], 
                                                      [pct_f1, pct_f2, pct_f3],
                                                      ['f1 (Delay) - xe.h', 'f2 (Queue) - xe', 'f3 (Stops) - lượt/h'])):
        ax.bar(nodes_x - 0.2, bl, 0.4, label='Baseline', color=COLORS[1], alpha=0.8)
        ax.bar(nodes_x + 0.2, opt, 0.4, label='Tối ưu', color=COLORS[0], alpha=0.8)
        ax.set_title(title, fontweight='bold')
        ax.set_xticks(nodes_x)
        ax.set_xticklabels(labels_short, rotation=45, fontsize=8)
        ax.legend(loc='upper left')
        
        ax_twin = ax.twinx()
        ax_twin.plot(nodes_x, pct, 'g^-', label='% Cải thiện', ms=6)
        ax_twin.axhline(0, color='green', linestyle=':', alpha=0.5)
        ax_twin.legend(loc='upper right')
        
    plt.tight_layout()
    plt.savefig('outputs/06_comparison.png')
    plt.close()
    return f1_opt, f2_opt, f3_opt, pct_f1, pct_f2, pct_f3

def build_result_table(best_ind, f1_opt, f2_opt, f3_opt, pct_f1, pct_f2, pct_f3):
    n = len(df_nodes)
    G1_opt, OFF_opt, c_opt = best_ind[:n], best_ind[n:2*n], best_ind[-1]
    G2_opt = [max(G_MIN, c_opt - G1_opt[i] - 2*int(df_nodes['L'][i])) for i in range(n)]

    data = {
        'Nút': list(range(1, n+1)),
        'Tên nút': df_nodes['name'].tolist(),
        'q (xe/h)': df_nodes['q'].round(1).tolist(),
        'C_BL (s)': BASELINE_C, 'C_OPT (s)': [c_opt] * n,
        'g1_BL (s)': BASELINE_G1, 'g1_OPT (s)': G1_opt,
        'g2_BL (s)': BASELINE_G2, 'g2_OPT (s)': G2_opt,
        'Off_BL (s)': BASELINE_OFF, 'Off_OPT (s)': OFF_opt,
        'f1_BL': BASELINE_F1, 'f1_OPT': f1_opt, 'Δf1 (%)': pct_f1,
        'f2_BL': BASELINE_F2, 'f2_OPT': f2_opt, 'Δf2 (%)': pct_f2,
        'f3_BL': BASELINE_F3, 'f3_OPT': f3_opt, 'Δf3 (%)': pct_f3,
        # BỔ SUNG: Xuất thêm thông tin L_max để tiện so sánh
        'L_max (xe)': L_MAX_LIST
    }
    return pd.DataFrame(data)

def export_result_table(df_result, best_ind, runtime_info):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        
        ws1 = wb.active
        ws1.title = 'Kết quả tối ưu'
        header_fill = PatternFill('solid', start_color='1565C0', end_color='1565C0')
        header_font = Font(color='FFFFFF', bold=True)
        
        cols = list(df_result.columns)
        for j, col in enumerate(cols):
            cell = ws1.cell(row=1, column=j+1, value=col)
            cell.fill, cell.font = header_fill, header_font
            
        for i, row in df_result.iterrows():
            for j, val in enumerate(row):
                ws1.cell(row=i+2, column=j+1, value=val)
                
        ws2 = wb.create_sheet(title='Thông tin Thuật toán')
        ws2.append(['MỤC THÔNG TIN', 'GIÁ TRỊ QUAN TRẮC'])
        for col in range(1, 3):
            cell = ws2.cell(row=1, column=col)
            cell.fill, cell.font = header_fill, header_font

        ws2.append(['Chu kỳ chung tìm được (C)', f"{best_ind[-1]} giây"])
        ws2.append(['Kích thước quần thể NSGA-II', runtime_info['pop_size']])
        ws2.append(['Số thế hệ tiến hóa NSGA-II', runtime_info['n_gen']])
        ws2.append(['Số lượng nghiệm trên đường Pareto', runtime_info['pareto_size']])
        ws2.append(['Số vòng lặp Local Search', runtime_info['ls_iter']])
        ws2.append(['Thời gian xử lý NSGA-II (giây)', round(runtime_info['nsga2_time'], 2)])
        ws2.append(['Thời gian xử lý Local Search (giây)', round(runtime_info['ls_time'], 2)])
        ws2.append(['TỔNG THỜI GIAN CHẠY (giây)', round(runtime_info['total_time'], 2)])
                
        wb.save('outputs/ket_qua_toi_uu_mang_luoi.xlsx')
        print("✓ Đã lưu: ket_qua_toi_uu_mang_luoi.xlsx")
    except Exception as e:
        print(f"Lỗi xuất Excel: {e}")

# ===================== MAIN =====================
def main():
    print("=" * 70)
    print("  TỐI ƯU HÓA ĐÈN GIAO THÔNG MẠNG LƯỚI – NSGA-II + LOCAL SEARCH")
    print("  CẬP NHẬT: ĐỒNG BỘ CHU KỲ CHUNG (C) CHO TOÀN BỘ MẠNG LƯỚI (26 NÚT)")
    print("  CẬP NHẬT: TÍCH HỢP HÀM PHẠT CHỐNG TRÀN BẾN L_MAX")
    print("=" * 70)

    t_start = time.time()

    load_data_from_excel()
    plot_input_data()

    weights, A, lambda_max, CI, CR = run_ahp_analysis()
    plot_ahp(A, weights, lambda_max, CI, CR)

    t_nsga = time.time()
    pareto_pop, pareto_fits, history = nsga2(pop_size=120, n_gen=80, seed=42)
    nsga2_time = time.time() - t_nsga

    t_ls = time.time()
    best_ind, best_fit, ls_history = local_search(pareto_pop, pareto_fits, weights, n_iter=60)
    ls_time = time.time() - t_ls

    t_total = time.time() - t_start
    runtime_info = {'pop_size': 120, 'n_gen': 80, 'ls_iter': 60, 'pareto_size': len(pareto_pop),
                    'nsga2_time': nsga2_time, 'ls_time': ls_time, 'total_time': t_total}

    print("\n[5] Tạo biểu đồ kết quả...")
    plot_algorithm_efficiency(history, ls_history)
    plot_pareto_front(pareto_fits, best_fit)
    plot_signal_plan(best_ind)
    f1_opt, f2_opt, f3_opt, pct_f1, pct_f2, pct_f3 = plot_comparison(best_ind, best_fit)
    
    print("\n[6] Xuất bảng kết quả...")
    df_result = build_result_table(best_ind, f1_opt, f2_opt, f3_opt, pct_f1, pct_f2, pct_f3)
    export_result_table(df_result, best_ind, runtime_info)

    print("\n" + "="*70 + "\n  KẾT QUẢ TỔNG KẾT MẠNG LƯỚI (Chu kỳ chung C = {}s)\n".format(best_ind[-1]) + "="*70)
    
    f1_bl_sum, f1_opt_sum = sum(BASELINE_F1), sum(f1_opt)
    f2_bl_sum, f2_opt_sum = sum(BASELINE_F2), sum(f2_opt)
    f3_bl_sum, f3_opt_sum = sum(BASELINE_F3), sum(f3_opt)
    
    print(f"  f1 Baseline   : {f1_bl_sum:.2f} xe.h")
    print(f"  f1 Tối ưu     : {f1_opt_sum:.2f} xe.h  → Cải thiện {(f1_bl_sum-f1_opt_sum)/f1_bl_sum*100:.1f}%")
    print(f"  f2 Baseline   : {f2_bl_sum:.2f} xe")  
    print(f"  f2 Tối ưu     : {f2_opt_sum:.2f} xe  → Cải thiện {(f2_bl_sum-f2_opt_sum)/f2_bl_sum*100:.1f}%") 
    print(f"  f3 Baseline   : {f3_bl_sum:.0f} lượt/h")
    print(f"  f3 Tối ưu     : {f3_opt_sum:.0f} lượt/h  → Cải thiện {(f3_bl_sum-f3_opt_sum)/f3_bl_sum*100:.1f}%")
    print(f"\n  Tổng thời gian: {t_total:.1f} giây")

if __name__ == '__main__':
    main()